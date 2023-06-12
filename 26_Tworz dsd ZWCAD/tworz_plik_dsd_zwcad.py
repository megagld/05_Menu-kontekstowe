import os
from openpyxl import load_workbook
from pathlib import Path

head=r'''[DWF6Version]
Ver=1
[DWF6MinorVersion]
MinorVer=1
'''

tail=r'''[Target]
Type=2
DWF=C:\Users\A2\Documents\0143_PAB_M_02_WD-2_02_420x0750
OUT=C:\Users\A2\Documents
PWD=
[PdfOptions]
CreateBookmarks=FALSE
[SheetSet Properties]
IsSheetSet=FALSE
IsHomogeneous=FALSE
SheetSet Name=
NoOfCopies=1
PlotStampOn=FALSE
JobID=0
SelectionSetName=
ZwcadProfile=
CategoryName=
LogFilePath=
IncludeLayer=FALSE
PromptForDwfName=FALSE
PwdProtectPublishedDWF=FALSE
PromptForPwd=FALSE
RepublishingMarkups=FALSE'''

def pobierz_dane(x):
    file = '{}\\{}'.format(input_dir,x)

    # Nazwy arkuszy:
    wb=load_workbook(file)
  
    ark=wb["DANE"]

    number_of_rows = ark.max_row
    last_row_index_with_data = 0
    
    while True:
        if ark.cell(number_of_rows, 3).value != None:
            last_row_index_with_data = number_of_rows
            break
        else:
            number_of_rows -= 1

    print(last_row_index_with_data)

    dane=[]
    for i in range(4,last_row_index_with_data+1):
        pat=ark.cell(row=i,column=2).value
        nam=ark.cell(row=i,column=3).value
        plot=ark.cell(row=i,column=44).value
        dane.append([pat,nam,plot])
    return dane

input_dir = os.getcwd()

def tworz_dsd(x):
    global input_dir, tail, head
    dane=pobierz_dane(x)  

    with open("{}\{}_zwcad.dsd".format(input_dir,x.split('.')[0]), 'w', encoding="utf-8") as file:
        file.write(head)

    for i,j,k in dane:
        if k=='x':
            inp=r'''[DWF6Sheet:{}]
        DWG={}
        Layout={}
        Setup={}
        OriginalSheetPath={}
        Has Plot Port=0
        Has3DDWF=0
        '''.format(j,i,j,'do PDF_'+j[-8:],i)

            with open("{}\{}_zwcad.dsd".format(input_dir,x.split('.')[0]), 'a', encoding="utf-8") as file:
                file.write(inp)

    with open("{}\{}_zwcad.dsd".format(input_dir,x.split('.')[0]), 'a', encoding="utf-8") as file:
        file.write(tail)


for x in os.listdir(input_dir):
    if not x.endswith('.xlsm'):
        continue
    else:
        try:
            tworz_dsd(x)
        except:
            pass